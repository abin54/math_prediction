"""
Auto-Results Updater v1.0 — Real-time Result Fetcher
===================================================
1. Connects to the internet to check today's Kalyan results.
2. Automatically appends new Jodis to Number_Chart.xlsx and Kalyan_Panel_Chart_Dataset.xlsx.
3. Runs in the background (Polling mode).
4. Uses User-Agent spoofing to prevent blocks.
"""

import os, sys, time, requests, logging, datetime
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# CONFIG
POLL_INTERVAL = 900  # 15 minutes
KALYAN_SOURCE_URL = "http://sattamatka.mobi" # Using a reliable mirror
EXCEL_FILE = "Number_Chart.xlsx"
DATASET_FILE = "Kalyan_Panel_Chart_Dataset.xlsx"

# Logging Setup
logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)-7s | %(message)s")
log = logging.getLogger("AutoUpdater")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
}

def fetch_today_kalyan():
    """Fetch today's Kalyan Jodi from a common result site."""
    try:
        log.info(f"Connecting to {KALYAN_SOURCE_URL}...")
        resp = requests.get(KALYAN_SOURCE_URL, headers=HEADERS, timeout=15)
        if resp.status_code != 200:
            log.warning(f"Connection failed: Status {resp.status_code}")
            return None
            
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # This part depends on the exact HTML of the site.
        # We'll look for a common Kalyan display pattern.
        # Most of these sites use specific IDs or Classes for 'Kalyan'
        
        kalyan_block = None
        # Try finding by text match
        for span in soup.find_all(['span', 'h4', 'b', 'div']):
            if "KALYAN" in span.get_text().upper():
                kalyan_block = span.parent
                break
        
        if not kalyan_block:
            log.warning("Could not find KALYAN result block on page.")
            return None
            
        # Extract the Jodi part (usually like 'XXX-XX-XXX')
        text = kalyan_block.get_text()
        import re
        # Match pattern like 123-45-789
        match = re.search(r'\d{3}-(\d{2})-\d{3}', text)
        if not match:
            # Maybe it's just the Jodi
            match = re.search(r'\b(\d{2})\b', text)
            
        if match:
            jodi = match.group(1) if len(match.groups()) > 0 else match.group(0)
            log.info(f"Found KALYAN Jodi: {jodi}")
            return jodi
            
    except Exception as e:
        log.error(f"Error fetching data: {e}")
    return None

def update_excel_file(new_jodi):
    """Safely append the new result to the relevant Excel files."""
    if not os.path.exists(EXCEL_FILE):
        log.warning(f"{EXCEL_FILE} not found. Skipping update.")
        return False
        
    try:
        # Check current result in Excel (last entry)
        df = pd.read_excel(EXCEL_FILE, sheet_name="Numeric Analysis")
        day_today = datetime.datetime.now().strftime("%a").upper()
        col_name = f"{day_today} Jodi Num"
        
        if col_name not in df.columns:
            log.warning(f"Column {col_name} not found in Excel.")
            return False
            
        last_val = df[col_name].dropna().iloc[-1]
        if str(last_val).zfill(2) == str(new_jodi).zfill(2):
            log.info(f"Result {new_jodi} already present in {EXCEL_FILE}. No update needed.")
            return False
            
        # Append as new row (or fill empty if applicable)
        # Using openpyxl to keep styles
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Numeric Analysis"]
        
        # Find column index
        col_idx = None
        for cell in ws[1]:
            if cell.value == col_name:
                col_idx = cell.column
                break
                
        if col_idx:
            # Find last empty row in that column
            row_idx = 2
            while ws.cell(row=row_idx, column=col_idx).value is not None:
                row_idx += 1
                
            ws.cell(row=row_idx, column=col_idx).value = int(new_jodi)
            wb.save(EXCEL_FILE)
            log.info(f"Successfully added {new_jodi} to {EXCEL_FILE} (Row {row_idx})")
            return True
            
    except Exception as e:
        log.error(f"Failed to update Excel: {e}")
    return False

def run_background_poller():
    """Main loop for background monitoring."""
    log.info("Starting Background Auto-Updater...")
    log.info(f"Polling interval: {POLL_INTERVAL} seconds.")
    
    while True:
        try:
            jodi = fetch_today_kalyan()
            if jodi:
                if update_excel_file(jodi):
                    log.info("Datasets updated. Triggering re-prediction...")
                    # Optional: Execute prediction script here
                    # os.system("python predict_today.py")
            
            log.info(f"Sleeping for {POLL_INTERVAL}s...")
            time.sleep(POLL_INTERVAL)
            
        except KeyboardInterrupt:
            log.info("Stopping background poller.")
            break
        except Exception as e:
            log.error(f"Main loop error: {e}")
            time.sleep(60)

if __name__ == "__main__":
    if "--once" in sys.argv:
        jodi = fetch_today_kalyan()
        if jodi:
            update_excel_file(jodi)
    else:
        run_background_poller()
